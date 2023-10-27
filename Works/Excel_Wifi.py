from bs4 import BeautifulSoup
import urllib.request
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time
import openpyxl as op


url = "https://www.localdata.go.kr/lif/lifeCtacDataView.do"

CHROME_DRIVER_PATH = './chromedriver.exe'
service = Service(executable_path=CHROME_DRIVER_PATH)
options = webdriver.ChromeOptions()
browser = webdriver.Chrome(service=service, options=options)

browser.get(url)
browser.implicitly_wait(10) # 로딩 대기

browser.execute_script("javascript:fDataView('12_04_07_E')")

search_city = Select(browser.find_element(By.CSS_SELECTOR, 'select#sidoNm'))
city = search_city.select_by_value('6270000')

search_button = browser.find_element(By.CSS_SELECTOR, 'a#searchBtn')
search_button.click()

time.sleep(3)

result = []
page = 1
while True:
    try:
        browser.execute_script(f'javascript:goPage({page})')
        page += 1
        time.sleep(1)
        
        
        html = browser.page_source
        soup = BeautifulSoup(html, 'html.parser')
        trs = soup.select('tr.etctablemodal')
        for tr in trs:
            td = tr.find_all('td')
            if len(td) < 6:
                break 
            install_name = td[1].string
            install_addr = td[2].string
            install_category = td[3].string
            install_company = td[4].string

            result.append([install_name, install_addr, install_category, install_company])

        else:
            continue
        break
        
    except:
        pass

print(result)

################################

import openpyxl as op

wb = op.Workbook()

ws = wb.create_sheet("wifi_list", 0)

col_names = ['번호', '설치장소명', '주소', '설치시설구분', '관리기관명']
for col, name in enumerate(col_names):
    ws.cell(row=1, column=col+1).value = name

for i, data in enumerate(result):
    ws.cell(row=i+2, column=1).value = i+1
    ws.cell(row=i+2, column=2).value = data[0]
    ws.cell(row=i+2, column=3).value = data[1]
    ws.cell(row=i+2, column=4).value = data[2]
    ws.cell(row=i+2, column=5).value = data[3]



wb.save("wifi.xlsx")