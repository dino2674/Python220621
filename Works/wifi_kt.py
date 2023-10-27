import pymysql

conn = pymysql.connect(host='localhost', 
                       user='admin', 
                       password='1111', 
                       db='rpa', 
                       charset='utf8')

cursor = conn.cursor()
query = "select * from wifi where install_company = 'KT'"
cursor.execute(query)

rows = cursor.fetchall()
print(rows[0])

import openpyxl as op

wb = op.load_workbook("./wifi.xlsx")

ws = wb.create_sheet("KT_LIST", 1)
ws['A1'] = "번호"
ws['B1'] = "설치장소"
ws['C1'] = "주소"
ws['D1'] = "설치시설"

for i, data in enumerate(rows):
    ws.cell(row=i+2, column=1).value = i+1
    ws.cell(row=i+2, column=2).value = data[1]
    ws.cell(row=i+2, column=3).value = data[2]
    ws.cell(row=i+2, column=4).value = data[3]

wb.save("wifi.xlsx")